"""XLSX export service — produces 23-column workbook matching golden contract."""

import io

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from db_models import Job, Tender

# Exact column order — must never change
COLUMNS = [
    "S.No", "Keyword", "Tender Reference No", "Tender Type",
    "Published Date", "Bid Submission End Date", "Title", "Description",
    "Cleaned BOQ", "Organisation", "Ministry", "Tender Value", "EMD",
    "State", "Pincode", "Delivery Period", "Product Type", "Exemption",
    "Email", "IT Relevant", "Quantity", "Link", "Scraped Date",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


async def generate_xlsx(db: AsyncSession, job_id: int, tenant_id: str) -> bytes:
    """Generate XLSX bytes for all tenders in a job.

    Verifies the job belongs to user_id before proceeding.
    Returns raw bytes suitable for streaming as a download response.
    """
    # Verify ownership
    job_result = await db.execute(
        select(Job).where(Job.id == job_id, Job.tenant_id == tenant_id)
    )
    if job_result.scalar_one_or_none() is None:
        from fastapi import HTTPException, status as http_status
        raise HTTPException(status_code=http_status.HTTP_404_NOT_FOUND, detail="Job not found")

    # Load tenders ordered by id (preserves insertion order)
    tenders_result = await db.execute(
        select(Tender).where(Tender.job_id == job_id).order_by(Tender.id)
    )
    tenders = tenders_result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tenders"

    # Header row
    ws.append(COLUMNS)
    for col_idx, _ in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Data rows
    for row_num, tender in enumerate(tenders, start=1):
        row_data = tender.to_export_row(row_num)
        ws.append([row_data[col] for col in COLUMNS])

    # Auto-fit column widths (cap at 60)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
