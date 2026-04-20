"""Tender export service — produces Excel file with curated columns."""

import io
from datetime import datetime, timezone
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from db_models import Tender


# Exact column order as per requirements
COLUMNS = [
    "Bid Number",
    "Bid Title",
    "Organization",
    "Ministry",
    "State",
    "Bid End Date",
    "Published Date",
    "Bid Value",
    "EMD Amount",
    "Category",
    "Keywords",
    "GeM URL",
    "Contact Details",
    "Status",
    "Extraction Date",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _format_date(dt) -> str:
    """Format date/datetime as DD-MM-YYYY."""
    if not dt:
        return "N/A"
    try:
        if hasattr(dt, "strftime"):
            return dt.strftime("%d-%m-%Y")
        return "N/A"
    except Exception:
        return "N/A"


def _format_currency(value) -> object:
    """Keep numeric fields as numbers only."""
    if value is None or value == "":
        return "N/A"
    try:
        return float(value)
    except (ValueError, TypeError):
        return "N/A"


def _safe_value(value) -> str:
    """Fill missing values with N/A."""
    if value is None or value == "":
        return "N/A"
    return str(value).strip()


async def generate_tender_xlsx(
    db: AsyncSession, tenant_id: str, search_term: Optional[str] = None
) -> bytes:
    """Generate XLSX bytes for all tenders (deduplicated and formatted).

    Args:
        db: Database session
        tenant_id: Current tenant ID
        search_term: Optional search term to filter tenders

    Returns:
        Raw bytes suitable for streaming as a download response
    """
    # Load ALL tenders for this tenant
    query = select(Tender).where(Tender.tenant_id == tenant_id)

    if search_term:
        # Apply search filter if provided
        pattern = f"%{search_term}%"
        from sqlalchemy import or_

        query = query.where(
            or_(
                Tender.title.ilike(pattern),
                Tender.organisation.ilike(pattern),
                Tender.keyword.ilike(pattern),
            )
        )

    # Order by bid_end_date ascending
    query = query.order_by(Tender.bid_end_date.asc())
    tenders_result = await db.execute(query)
    tenders = tenders_result.scalars().all()

    # Remove duplicates by tender_ref_no, keeping first occurrence
    seen = set()
    unique_tenders = []
    for tender in tenders:
        if tender.tender_ref_no not in seen:
            seen.add(tender.tender_ref_no)
            unique_tenders.append(tender)

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GeM_Tender_Data"

    # Header row
    ws.append(COLUMNS)
    for col_idx, _ in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)

    # Data rows
    for tender in unique_tenders:
        row_data = [
            _safe_value(tender.tender_ref_no),  # Bid Number
            _safe_value(tender.title),  # Bid Title
            _safe_value(tender.organisation),  # Organization
            _safe_value(tender.ministry),  # Ministry
            _safe_value(tender.state),  # State
            _format_date(tender.bid_end_date),  # Bid End Date
            _format_date(tender.published_date),  # Published Date
            _format_currency(tender.tender_value),  # Bid Value
            _format_currency(tender.emd),  # EMD Amount
            _safe_value(tender.product_type),  # Category
            _safe_value(tender.keyword),  # Keywords
            _safe_value(tender.link),  # GeM URL
            _safe_value(tender.email),  # Contact Details
            "Active",  # Status (all are active by default)
            _format_date(tender.scraped_date),  # Extraction Date
        ]
        ws.append(row_data)

    # Auto-fit column widths (cap at 60)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    # Enable filter on all columns
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    # Freeze top row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
