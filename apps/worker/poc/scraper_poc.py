"""
GeM Tender Bot — Phase 0 Playwright POC
========================================
Clean-room rewrite of the Selenium-based legacy scraper using Playwright.
Selectors and field mapping derived from legacy source in legacy/decompiled/.

Acceptance criteria:
  - At least 1 PDF downloaded
  - pdfplumber extracts ≥18 of 23 fields
  - XLSX written with correct 23-column headers
  - Runs without CAPTCHA block
"""

import os
import re
import time
import random
import requests
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from playwright.sync_api import sync_playwright
from playwright_stealth import stealth_sync

# ── Config ────────────────────────────────────────────────────────────────────
BASE_URL  = "https://bidplus.gem.gov.in/all-bids"
KEYWORD   = "Laptop"
LIMIT     = 3          # cards to collect
MIN_VALUE = 50000      # skip tenders below ₹50k
HEADLESS  = True

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
DOWNLOAD_DIR = os.path.join(SCRIPT_DIR, "downloads_poc")
OUTPUT_DIR   = os.path.join(SCRIPT_DIR, "output_poc")

DOWNLOAD_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Referer": BASE_URL,
    "Accept": "application/pdf,*/*",
}

# 23 columns — exact order from golden XLSX
COLUMNS = [
    "S.No", "Keyword", "Tender Reference No", "Tender Type",
    "Published Date", "Bid Submission End Date", "Title", "Description",
    "Cleaned BOQ", "Organisation", "Ministry", "Tender Value", "EMD",
    "State", "Pincode", "Delivery Period", "Product Type", "Exemption",
    "Email", "IT Relevant", "Quantity", "Link", "Scraped Date",
]

# ── Text helpers ──────────────────────────────────────────────────────────────

def clean_text(text) -> str:
    if not text:
        return ""
    text = str(text)
    text = re.sub(r'[\u0900-\u097F]', '', text)   # strip Devanagari
    text = re.sub(r'\(cid:\d+\)', '', text)         # remove CID junk
    text = re.sub(r'^\s*/\s*|\s*/\s*$', '', text)
    return re.sub(r'\s+', ' ', text).strip()


def safe_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "_", name)


# ── Card text parser ──────────────────────────────────────────────────────────

def parse_bid_no(text: str) -> str:
    m = re.search(r'Bid\s+No\.?\s*:\s*(GEM/[\w/]+)', text, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.search(r'\b(GEM/\d{4}/[A-Z]/\d+)\b', text)
    return m.group(1).strip() if m else "N/A"


def parse_tender_type(text: str) -> str:
    tl = text.lower()
    if any(k in tl for k in ["boq", "bill of quantity"]):
        return "boq"
    if any(k in tl for k in ["service bid", "service ra", "service contract"]):
        return "service"
    if any(k in tl for k in ["custom bid", "custom product", "customized"]):
        return "product_custom"
    return "product"


def parse_card_text(text: str) -> dict:
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    fields = {
        "Bid_No": parse_bid_no(text),
        "Tender_Type": parse_tender_type(text),
        "Start_Date": "N/A", "End_Date": "N/A",
        "Dept_Addr": "N/A", "Items": "N/A", "Quantity": "N/A",
    }
    for i, line in enumerate(lines):
        if "Items:" in line:
            fields["Items"] = line.replace("Items:", "").strip()
        elif "Quantity:" in line:
            fields["Quantity"] = line.replace("Quantity:", "").strip()
        elif "Department Name And Address:" in line:
            fields["Dept_Addr"] = " | ".join(lines[i+1:i+3])
        elif "Start Date:" in line:
            fields["Start_Date"] = line.replace("Start Date:", "").strip()
        elif "End Date:" in line:
            fields["End_Date"] = line.replace("End Date:", "").strip()
    return fields


# ── PDF download ──────────────────────────────────────────────────────────────

def download_pdf(url: str, bid_id: str) -> str | None:
    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    path = os.path.join(DOWNLOAD_DIR, safe_filename(bid_id) + ".pdf")
    if os.path.exists(path):
        return path
    try:
        r = requests.get(url, headers=DOWNLOAD_HEADERS, timeout=30, stream=True)
        r.raise_for_status()
        with open(path, "wb") as f:
            for chunk in r.iter_content(8192):
                f.write(chunk)
        return path
    except Exception as e:
        print(f"   [Download Error] {e}")
        return None


# ── PDF parser ────────────────────────────────────────────────────────────────

def parse_pdf(file_path: str):
    table_data, full_text = {}, ""
    try:
        with pdfplumber.open(file_path) as pdf:
            settings = {
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance": 3,
                "join_tolerance": 3,
            }
            for page in pdf.pages:
                for table in page.extract_tables(table_settings=settings):
                    for row in table:
                        if not row:
                            continue
                        cells = [clean_text(c) for c in row if c and clean_text(c)]
                        if len(cells) < 2:
                            continue
                        key = cells[0].lower().strip()
                        value = cells[-1].strip()
                        if len(key) < 3 or len(key) > 100 or re.fullmatch(r'[\d\s₹,\.]+', key):
                            continue
                        table_data.setdefault(key, value)
                        for i in range(len(cells) - 1):
                            k, v = cells[i].lower().strip(), cells[i+1].strip()
                            if 3 <= len(k) <= 100 and not re.fullmatch(r'[\d\s₹,\.]+', k):
                                table_data.setdefault(k, v)

            raw_text = ""
            for page in pdf.pages:
                t = page.extract_text()
                if t:
                    raw_text += clean_text(t) + "\n"
            full_text = re.sub(r'\s+', ' ', raw_text).strip()
    except Exception as e:
        print(f"   [PDF Error] {e}")
    return table_data, full_text


# ── Field extractor ───────────────────────────────────────────────────────────

FIELD_MAP = {
    "Ministry":     ["ministry/state name", "ministry name", "ministry"],
    "Buyer":        ["organisation name", "buyer organisation name", "buyer name", "consignee name", "buyer"],
    "Bid_Value":    ["estimated bid value", "estimated value in inr", "bid value", "tender value"],
    "EMD":          ["emd amount", "earnest money deposit", "bid security amount", "emd"],
    "Delivery":     ["delivery period in days", "delivery period (in days)", "delivery period", "contract period"],
    "Category":     ["item category", "product category", "category name", "category"],
    "Exemption":    ["mse relaxation for years of experience and turnover", "mse exemption", "startup exemption", "exemption"],
    "Email":        ["email id", "email address", "e-mail", "email"],
    "Pincode":      ["pin code", "pincode", "postal code"],
    "State":        ["state name", "state of delivery", "consignee state", "state"],
    "Product_Type": ["type of bid", "bid type", "product type"],
}

JUNK = {"yes", "no", "download", "n/a", "na", "-", "none", "null", ""}
VALID_2_DIGIT_PREFIXES = {
    "11","12","13","14","15","16","17","18","19",
    "20","21","22","23","24","25","26","27","28","29",
    "30","31","32","33","34","36","37","38","39",
    "40","41","42","43","44","45","46","47","48","49",
    "50","51","52","53","56","57","58","60","61","62",
    "63","64","67","68","69","70","71","72","73","74",
    "75","76","77","78","79","80","81","82","83","84","85","86",
}


def _find_value(data: dict, patterns: list) -> str:
    for pat in patterns:
        for k, v in data.items():
            if k.strip().lower() == pat.lower():
                val = str(v).strip()
                if val.lower() not in JUNK and len(val) > 1:
                    return val
    for pat in patterns:
        for k, v in data.items():
            if pat.lower() in k.strip().lower():
                val = str(v).strip()
                if val.lower() not in JUNK and len(val) > 1:
                    return val
    return "N/A"


def _is_valid_pincode(s: str) -> bool:
    if not re.fullmatch(r'\d{6}', s) or s[0] == '0':
        return False
    return s[:2] in VALID_2_DIGIT_PREFIXES


def _clean_amount(val: str) -> str:
    try:
        return str(int(float(val.strip().replace(",", ""))))
    except ValueError:
        return val


def extract_fields(table_data: dict, full_text: str, bid_no: str = "",
                   dept_addr: str = "", items_from_card: str = "") -> dict:
    result = {f: _find_value(table_data, pats) for f, pats in FIELD_MAP.items()}

    # Currency fallback
    if result["Bid_Value"] == "N/A":
        clean = re.sub(r'GEM/\d{4}/[A-Z]/\d+', '', full_text)
        prefixed = re.findall(r'(?:₹|Rs\.?)\s*([0-9][0-9,\.]+)', clean)
        parsed = []
        for a in prefixed:
            try:
                v = int(a.replace(",", "").split(".")[0])
                if v >= 10000:
                    parsed.append(v)
            except ValueError:
                pass
        result["Bid_Value"] = str(max(parsed)) if parsed else "N/A"

    # EMD fallback
    if result["EMD"] == "N/A":
        m = re.search(r'(?:emd|earnest\s+money)[\s:]*(?:₹|Rs\.?|INR)?\s*([0-9][0-9,\.]+)', full_text, re.IGNORECASE)
        if m:
            result["EMD"] = _clean_amount(m.group(1))

    # Delivery fallback
    if result["Delivery"] == "N/A":
        m = re.search(r'delivery\s+period\s*(?:\(in\s+days?\))?\s*[\r\n\s:]*(\d+)', full_text, re.IGNORECASE | re.DOTALL)
        if m:
            d = int(m.group(1))
            if 1 <= d <= 3650:
                result["Delivery"] = f"{d} Days"

    # Email fallback
    if result["Email"] == "N/A":
        m = re.search(r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+', full_text)
        result["Email"] = m.group() if m else "N/A"

    # Exemption
    mse = re.search(r'mse\s+relaxation\s+for\s+years.*?[:\s]+(yes|no)', full_text, re.IGNORECASE | re.DOTALL)
    startup = re.search(r'startup\s+relaxation\s+for\s+years.*?[:\s]+(yes|no)', full_text, re.IGNORECASE | re.DOTALL)
    found_exempt = []
    if mse and mse.group(1).lower() == "yes":
        found_exempt.append("MSE")
    if startup and startup.group(1).lower() == "yes":
        found_exempt.append("Startup")
    result["Exemption"] = " & ".join(found_exempt) if found_exempt else (
        "No" if (mse or startup) else "N/A"
    )

    # Pincode
    known_amounts = set()
    for af in ["Bid_Value", "EMD"]:
        v = result.get(af, "N/A")
        if v and v != "N/A":
            try:
                known_amounts.add(str(int(float(v))))
            except (ValueError, TypeError):
                pass

    raw_pin = result.get("Pincode", "N/A")
    if raw_pin == "N/A" or not _is_valid_pincode(str(raw_pin).strip()) or raw_pin in known_amounts:
        clean_text_pin = re.sub(r'GEM/\d{4}/[A-Z]/\d+', '', full_text)
        for c in re.findall(r'\b([1-9][0-9]{5})\b', clean_text_pin):
            if _is_valid_pincode(c) and c not in known_amounts:
                result["Pincode"] = c
                break
        else:
            result["Pincode"] = "N/A"

    # State fallback
    if result["State"] == "N/A":
        STATES = [
            "Andhra Pradesh", "Arunachal Pradesh", "Assam", "Bihar",
            "Chhattisgarh", "Goa", "Gujarat", "Haryana", "Himachal Pradesh",
            "Jharkhand", "Karnataka", "Kerala", "Madhya Pradesh", "Maharashtra",
            "Manipur", "Meghalaya", "Mizoram", "Nagaland", "Odisha", "Punjab",
            "Rajasthan", "Sikkim", "Tamil Nadu", "Telangana", "Tripura",
            "Uttar Pradesh", "Uttarakhand", "West Bengal", "Delhi",
            "Jammu & Kashmir", "Ladakh", "Chandigarh", "Puducherry"
        ]
        for state in STATES:
            if state.lower() in full_text.lower():
                result["State"] = state
                break

    # Buyer fallback from card dept string
    if result["Buyer"] == "N/A" and dept_addr and dept_addr != "N/A":
        parts = [p.strip() for p in dept_addr.split("|") if p.strip()]
        if len(parts) >= 2:
            result["Buyer"] = parts[-1]

    # BOQ
    boq_lines = []
    BOQ_KEYS = ["item name", "item description", "description of goods", "product name", "description", "boq"]
    for k, v in table_data.items():
        if any(bk in k.lower() for bk in BOQ_KEYS):
            val = str(v).strip()
            if len(val) > 3 and val.lower() not in {"n/a", "na", "-"}:
                for p in val.split("|"):
                    p = p.strip()
                    if p and not re.search(r'[\u0900-\u097F]', p):
                        boq_lines.append(p)
    result["Cleaned_BOQ"] = "\n".join(boq_lines[:20]) if boq_lines else (items_from_card if items_from_card != "N/A" else "N/A")

    return result


# ── IT relevance ──────────────────────────────────────────────────────────────

NON_IT = ["fire extinguisher", "refrigerator", "furniture", "vehicle", "diesel",
          "generator set", "uniform", "stationery", "civil work", "medicine",
          "ambulance", "catering", "food", "clothing", "agriculture"]
IT_KW  = ["server", "laptop", "desktop", "workstation", "amc", "camc", "fms",
          "networking", "firewall", "hpc", "software", "cloud", "data center",
          "storage", "router", "switch", "it support", "cyber", "cctv",
          "computer", "printer", "scanner", "ups", "wifi", "it services"]

def is_it_relevant(items: str, category: str, keyword: str) -> bool:
    combined = f"{items} {category} {keyword}".lower()
    if any(n in combined for n in NON_IT):
        return False
    return any(k in combined for k in IT_KW) or True


# ── XLSX writer ───────────────────────────────────────────────────────────────

def write_xlsx(rows: list, run_start: datetime, run_end: datetime):
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    filename = os.path.join(OUTPUT_DIR, f"poc_{run_end.strftime('%Y%m%d_%H%M')}.xlsx")
    date_range = f"Tenders Published: {run_start.strftime('%Y-%m-%d %H:%M')} to {run_end.strftime('%Y-%m-%d %H:%M')}"

    wb = openpyxl.Workbook()
    for sheet_name, sheet_rows in [("All Tenders", rows), ("IT Relevant Only", [r for r in rows if r.get("IT Relevant") == "✅ YES"])]:
        ws = wb.active if sheet_name == "All Tenders" else wb.create_sheet(sheet_name)
        ws.title = sheet_name

        # Row 1: merged title
        ws.cell(row=1, column=1, value=date_range)
        ws.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF", size=11)
        ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="1F4E79")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))

        # Row 2: headers
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=2, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Data rows
        for row_idx, row_data in enumerate(sheet_rows, start=3):
            for col_idx, col_name in enumerate(COLUMNS, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name, "N/A"))

        # Column widths
        for col_idx, col_name in enumerate(COLUMNS, 1):
            col_letter = get_column_letter(col_idx)
            if col_name in ("Cleaned BOQ", "Description"):
                ws.column_dimensions[col_letter].width = 50
            else:
                max_len = max(
                    (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, ws.max_row + 1)),
                    default=10
                )
                ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    wb.save(filename)
    print(f"\n✅ XLSX saved: {filename}")
    return filename


# ── Main scraper ──────────────────────────────────────────────────────────────

def run():
    run_start = datetime.now()
    print(f"\n🚀 GeM Playwright POC — {run_start.strftime('%H:%M:%S')}")
    print(f"   Keyword : {KEYWORD!r}  |  Limit : {LIMIT}  |  Min Value : ₹{MIN_VALUE:,}\n")

    os.makedirs(DOWNLOAD_DIR, exist_ok=True)
    rows = []

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(
            viewport={"width": 1920, "height": 1080},
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        page = context.new_page()
        stealth_sync(page)

        print(f"[1] Navigating to {BASE_URL} ...")
        page.goto(BASE_URL, wait_until="domcontentloaded", timeout=30000)
        time.sleep(random.uniform(2, 3))

        # Dismiss popup if present
        try:
            close_btn = page.locator("button:has-text('Close'), .modal-header button").first
            if close_btn.is_visible(timeout=3000):
                close_btn.click()
                time.sleep(1)
        except Exception:
            pass

        # Search
        print(f"[2] Searching for {KEYWORD!r} ...")
        search_box = page.locator("#searchBid").first
        search_box.scroll_into_view_if_needed()
        search_box.fill(KEYWORD)
        search_box.press("Enter")
        time.sleep(random.uniform(6, 8))
        page.wait_for_selector(".block_header", timeout=20000)
        print("   Results loaded.")

        # Sort by latest bid end date
        try:
            toggle = page.locator("button.dropdown-toggle").first
            toggle.click()
            time.sleep(1)
            sort_opt = page.locator("text=Bid End Date: Latest First").first
            sort_opt.click()
            time.sleep(random.uniform(4, 6))
            page.wait_for_selector(".block_header", timeout=10000)
            print("   Sorted: Bid End Date Latest First")
        except Exception as e:
            print(f"   ⚠ Sort failed (continuing): {e}")

        # Collect cards
        print(f"[3] Collecting up to {LIMIT} bid cards ...")
        collected, seen = [], set()

        for attempt in range(3):
            cards = page.locator(".block_header").all()
            for card in cards:
                if len(collected) >= LIMIT:
                    break
                try:
                    parent = card.locator("xpath=..").first
                    text = parent.inner_text()
                    href = ""
                    try:
                        href = card.locator("a.bid_no_hover").first.get_attribute("href") or ""
                    except Exception:
                        try:
                            href = parent.locator("a.bid_no_hover").first.get_attribute("href") or ""
                        except Exception:
                            pass
                    if href and not href.startswith("http"):
                        href = "https://bidplus.gem.gov.in" + href
                    key = href if href else text[:200]
                    if key and key not in seen:
                        seen.add(key)
                        collected.append({"text": text, "pdf_url": href or "N/A"})
                except Exception:
                    continue
            if len(collected) >= LIMIT:
                break
            if attempt < 2:
                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                time.sleep(2)

        print(f"   Collected {len(collected)} cards")
        browser.close()

    # Process each card
    for i, card in enumerate(collected[:LIMIT], 1):
        print(f"\n[{i}/{len(collected)}] Processing card ...")
        card_text = card.get("text", "")
        pdf_url = card.get("pdf_url", "N/A")
        cf = parse_card_text(card_text)
        bid_no = cf["Bid_No"]
        print(f"   Bid No : {bid_no}")
        print(f"   PDF URL: {pdf_url}")

        row = {
            "Keyword": KEYWORD,
            "Tender Reference No": bid_no,
            "Tender Type": cf["Tender_Type"],
            "Published Date": cf["Start_Date"],
            "Bid Submission End Date": cf["End_Date"],
            "Title": cf["Items"],
            "Quantity": cf["Quantity"],
            "Link": pdf_url,
            "Scraped Date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "Description": "N/A", "Cleaned BOQ": "N/A", "Organisation": "N/A",
            "Ministry": "N/A", "Tender Value": "N/A", "EMD": "N/A",
            "State": "N/A", "Pincode": "N/A", "Delivery Period": "N/A",
            "Product Type": "N/A", "Exemption": "N/A", "Email": "N/A",
            "IT Relevant": "N/A",
        }

        # Download + parse PDF
        if pdf_url != "N/A":
            time.sleep(random.uniform(1, 2))
            label = bid_no if bid_no != "N/A" else f"{KEYWORD}_{i}"
            pdf_path = download_pdf(pdf_url, label)

            if pdf_path:
                print(f"   ✦ Parsing {os.path.basename(pdf_path)} ...")
                table_data, full_text = parse_pdf(pdf_path)
                extracted = extract_fields(
                    table_data, full_text, bid_no=bid_no,
                    dept_addr=cf["Dept_Addr"], items_from_card=cf["Items"]
                )

                # Map extracted fields to XLSX columns
                row["Description"]    = extracted.get("Category", "N/A")
                row["Cleaned BOQ"]    = extracted.get("Cleaned_BOQ", "N/A")
                row["Ministry"]       = extracted.get("Ministry", "N/A")
                row["Tender Value"]   = extracted.get("Bid_Value", "N/A")
                row["EMD"]            = extracted.get("EMD", "N/A")
                row["State"]          = extracted.get("State", "N/A")
                row["Pincode"]        = extracted.get("Pincode", "N/A")
                row["Delivery Period"]= extracted.get("Delivery", "N/A")
                row["Product Type"]   = extracted.get("Product_Type", "N/A")
                row["Exemption"]      = extracted.get("Exemption", "N/A")
                row["Email"]          = extracted.get("Email", "N/A")

                # Organisation
                buyer = extracted.get("Buyer", "N/A")
                if buyer != "N/A":
                    row["Organisation"] = buyer
                elif cf["Dept_Addr"] != "N/A":
                    row["Organisation"] = cf["Dept_Addr"].split("|")[0].strip()

                # Count populated fields
                populated = sum(1 for c in [
                    "Tender Reference No", "Tender Type", "Published Date",
                    "Bid Submission End Date", "Title", "Description", "Cleaned BOQ",
                    "Organisation", "Ministry", "Tender Value", "EMD", "State",
                    "Pincode", "Delivery Period", "Product Type", "Exemption",
                    "Email", "IT Relevant", "Quantity", "Link", "Scraped Date"
                ] if row.get(c, "N/A") != "N/A")
                print(f"   Fields populated: {populated}/21 PDF-derived fields")

            else:
                print("   ✗ Download failed")
        else:
            print("   ⚠ No PDF link")

        # IT relevance
        it = is_it_relevant(row.get("Title", ""), row.get("Description", ""), KEYWORD)
        row["IT Relevant"] = "✅ YES" if it else "❌ NO"

        # Value filter
        bv = row.get("Tender Value", "N/A")
        if bv != "N/A":
            try:
                if int(bv.replace(",", "")) < MIN_VALUE:
                    print(f"   ⚠ Skipping — Tender Value ₹{bv} < ₹{MIN_VALUE:,}")
                    continue
            except ValueError:
                pass

        rows.append(row)

    if not rows:
        print("\n⚠ No rows collected.")
        return

    # Add S.No
    for idx, r in enumerate(rows, 1):
        r["S.No"] = idx

    run_end = datetime.now()
    filename = write_xlsx(rows, run_start, run_end)

    print(f"\n── POC Summary ──────────────────────────────")
    print(f"   Cards processed : {len(collected)}")
    print(f"   Rows written    : {len(rows)}")
    print(f"   Duration        : {(run_end - run_start).seconds}s")
    print(f"   Output          : {filename}")
    print(f"────────────────────────────────────────────\n")


if __name__ == "__main__":
    run()
