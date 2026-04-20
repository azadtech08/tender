import argparse
import csv
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

SHEET_NAME = "GeM_Tender_Data"
ERROR_SHEET = "Errors"

COLUMNS = [
    "Bid Number",
    "Bid Title",
    "Organization / Department",
    "Ministry",
    "State",
    "Bid End Date",
    "Published Date",
    "Bid Value (INR)",
    "EMD Amount",
    "Category (Product/Service)",
    "Keywords Matched",
    "GeM URL",
    "Contact Details (if available)",
    "Status (Active/Closed)",
    "Extraction Date",
]

STATUS_COLORS = {
    "Active": PatternFill(fill_type="solid", fgColor="C6EFCE"),
    "Closed": PatternFill(fill_type="solid", fgColor="FFC7CE"),
}

DATE_PATTERNS = [
    "%d-%m-%Y",
    "%d/%m/%Y",
    "%d.%m.%Y",
    "%Y-%m-%d",
    "%Y/%m/%d",
    "%d %b %Y",
    "%d %B %Y",
    "%d %b, %Y",
    "%d %B, %Y",
    "%d %m %Y",
    "%d %b %y",
    "%d %B %y",
    "%Y%m%d",
]

NUMBER_PATTERN = re.compile(r"[^0-9.]|\u0007")

FIELD_MAP = {
    "bid_number": ["bid number", "bid no", "tender ref", "tender reference", "reference no", "reference", "ref no", "tender id"],
    "bid_title": ["bid title", "title", "tender title", "subject", "headline", "description"],
    "organization": ["organization", "organisation", "department", "dept", "agency", "office", "organisation / department"],
    "ministry": ["ministry", "authority", "division"],
    "state": ["state", "location", "place", "district"],
    "bid_end_date": ["bid end date", "submission end date", "due date", "last date", "end date", "closing date", "bid closing"],
    "published_date": ["published date", "publish date", "publication date", "published on", "date of publication", "posted on", "date"],
    "bid_value": ["bid value", "tender value", "estimated value", "value", "estimated bid value", "approximate value"],
    "emd_amount": ["emd amount", "emd", "earnest money", "earnest money deposit"],
    "category": ["category", "product/service", "tender type", "type", "procurement type"],
    "keywords": ["keywords matched", "keywords", "matched keywords", "search terms", "search keyword"],
    "gem_url": ["gem url", "url", "link", "gem link", "tender link", "page link"],
    "contact_details": ["contact details", "contact", "phone", "email", "contact person", "contact info", "mobile"],
    "status": ["status", "state", "active", "closed", "tender status"],
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (list, tuple, set)):
        return " · ".join(normalize_text(item) for item in value if item is not None)
    text = str(value).strip()
    text = re.sub(r"[\t\r\n]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text


def parse_date(value: Any) -> Tuple[str, Optional[datetime]]:
    text = normalize_text(value)
    if not text:
        return "N/A", None

    text = text.replace("st", "").replace("nd", "").replace("rd", "").replace("th", "")
    text = text.replace("/", "-").replace(".", "-").strip()
    text = re.sub(r"[^0-9A-Za-z\- ]+", " ", text).strip()

    for pattern in DATE_PATTERNS:
        try:
            parsed = datetime.strptime(text, pattern)
            return parsed.strftime("%d-%m-%Y"), parsed
        except ValueError:
            continue

    digits = re.sub(r"[^0-9]", "", text)
    if len(digits) == 8:
        try:
            parsed = datetime.strptime(digits, "%Y%m%d")
            return parsed.strftime("%d-%m-%Y"), parsed
        except ValueError:
            pass

    return "N/A", None


def parse_number(value: Any) -> Any:
    text = normalize_text(value)
    if not text:
        return "N/A"

    cleaned = NUMBER_PATTERN.sub("", text.replace(",", "").replace("₹", "").replace("Rs", "").replace("INR", "").strip())
    if cleaned.count(".") > 1:
        cleaned = cleaned.replace(".", "", cleaned.count(".") - 1)
    if not cleaned or cleaned == ".":
        return "N/A"

    try:
        number = float(cleaned)
        if number.is_integer():
            return int(number)
        return number
    except ValueError:
        return "N/A"


def normalize_status(value: Any) -> str:
    text = normalize_text(value).lower()
    if "closed" in text:
        return "Closed"
    if "active" in text:
        return "Active"
    if text in {"open", "live", "running"}:
        return "Active"
    if text in {"expired", "cancelled", "cancelled"}:
        return "Closed"
    return "N/A"


def normalize_category(value: Any) -> str:
    text = normalize_text(value).lower()
    if not text:
        return "N/A"
    if "product" in text or "goods" in text or "material" in text:
        return "Product"
    if "service" in text or "consult" in text or "support" in text:
        return "Service"
    return text.title()


def match_field(key: str) -> Optional[str]:
    normalized_key = re.sub(r"[^a-z0-9 ]+", " ", key.lower())
    for field_name, patterns in FIELD_MAP.items():
        for pattern in patterns:
            if pattern in normalized_key:
                return field_name
    return None


def build_record(raw: Any) -> Optional[Dict[str, Any]]:
    if isinstance(raw, str):
        data = parse_text_record(raw)
    elif isinstance(raw, dict):
        data = {normalize_text(k): v for k, v in raw.items()}
    else:
        return None

    row: Dict[str, Any] = {key: "N/A" for key in COLUMNS}
    row["Extraction Date"] = datetime.utcnow().strftime("%d-%m-%Y")

    for raw_key, raw_value in data.items():
        field = match_field(raw_key)
        value = normalize_text(raw_value)
        if not field:
            continue

        if field == "bid_number":
            row["Bid Number"] = value
        elif field == "bid_title":
            if row["Bid Title"] == "N/A" or len(value) > len(row["Bid Title"]):
                row["Bid Title"] = value
        elif field == "organization":
            row["Organization / Department"] = value
        elif field == "ministry":
            row["Ministry"] = value
        elif field == "state":
            row["State"] = value
        elif field == "bid_end_date":
            parsed, _ = parse_date(value)
            row["Bid End Date"] = parsed
        elif field == "published_date":
            parsed, _ = parse_date(value)
            row["Published Date"] = parsed
        elif field == "bid_value":
            row["Bid Value (INR)"] = parse_number(value)
        elif field == "emd_amount":
            row["EMD Amount"] = parse_number(value)
        elif field == "category":
            row["Category (Product/Service)"] = normalize_category(value)
        elif field == "keywords":
            row["Keywords Matched"] = value
        elif field == "gem_url":
            row["GeM URL"] = value
        elif field == "contact_details":
            row["Contact Details (if available)"] = value
        elif field == "status":
            row["Status (Active/Closed)"] = normalize_status(value)

    if row["Bid Number"] == "N/A" and row["Bid Title"] == "N/A":
        return None

    return row


def parse_text_record(text: str) -> Dict[str, str]:
    lines = [line.strip() for line in re.split(r"\r?\n", text) if line.strip()]
    record: Dict[str, str] = {}
    current_key: Optional[str] = None
    current_value: List[str] = []

    for line in lines:
        if ":" in line or "=" in line or " - " in line:
            if current_key is not None:
                record[current_key] = " ".join(current_value).strip()
            if ":" in line:
                parts = line.split(":", 1)
            elif "=" in line:
                parts = line.split("=", 1)
            else:
                parts = line.split(" - ", 1)
            current_key = parts[0].strip()
            current_value = [parts[1].strip()] if len(parts) > 1 else []
        else:
            if current_key is not None:
                current_value.append(line)
            elif "http" in line.lower():
                record["GeM URL"] = line.strip()
            else:
                inferred = infer_free_text(line)
                if inferred:
                    record.update(inferred)

    if current_key is not None:
        record[current_key] = " ".join(current_value).strip()

    return record


def infer_free_text(line: str) -> Dict[str, str]:
    if "http" in line.lower():
        return {"GeM URL": line.strip()}
    if "rs" in line.lower() and any(digit.isdigit() for digit in line):
        return {"Bid Value": line}
    if re.search(r"\b(em|emd)\b", line, re.IGNORECASE):
        return {"EMD Amount": line}
    return {}


def load_input(path: Path) -> List[Any]:
    text = path.read_text(encoding="utf-8", errors="ignore")
    suffix = path.suffix.lower()

    if suffix in {".json"}:
        data = json.loads(text)
        if isinstance(data, list):
            return data
        if isinstance(data, dict):
            for key in ("items", "data", "tenders", "records", "rows"):
                if isinstance(data.get(key), list):
                    return data[key]
            return [data]
        return []

    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        reader = csv.DictReader(text.splitlines(), delimiter=delimiter)
        return [dict(row) for row in reader]

    return [block for block in re.split(r"\n\s*\n", text) if block.strip()]


def write_workbook(rows: List[Dict[str, Any]], errors: List[Dict[str, Any]], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    for col_idx, header in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, header in enumerate(COLUMNS, start=1):
            value = row.get(header, "N/A")
            ws.cell(row=row_idx, column=col_idx, value=value)
            if header == "Status (Active/Closed)" and value in STATUS_COLORS:
                ws.cell(row=row_idx, column=col_idx).fill = STATUS_COLORS[value]

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{len(rows)+1}"
    ws.freeze_panes = "A2"

    for col_idx in range(1, len(COLUMNS) + 1):
        width = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(rows) + 2)
        ) + 4
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(width, 12), 60)

    if errors:
        err_sheet = wb.create_sheet(ERROR_SHEET)
        err_sheet.append(["Reason", "Raw Input"])
        err_sheet.freeze_panes = "A2"
        for cell in err_sheet[1]:
            cell.font = Font(bold=True)
        for error in errors:
            err_sheet.append([error.get("reason", "Unknown"), normalize_text(error.get("raw", ""))])
        err_sheet.column_dimensions["A"].width = 40
        err_sheet.column_dimensions["B"].width = 80

    os.makedirs(output_path.parent, exist_ok=True)
    wb.save(output_path)
    print(f"✅ Excel file written: {output_path}")


def transform_file(input_path: Path, output_path: Optional[Path] = None) -> Path:
    raw_items = load_input(input_path)
    records: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    seen_ids: set[str] = set()

    for raw in raw_items:
        row = build_record(raw)
        if row is None:
            errors.append({"reason": "Skipped incomplete entry", "raw": raw})
            continue

        bid_num = normalize_text(row["Bid Number"]) or "N/A"
        if bid_num == "N/A":
            errors.append({"reason": "Missing Bid Number", "raw": raw})
            continue

        if bid_num in seen_ids:
            continue
        seen_ids.add(bid_num)

        if row["Bid End Date"] == "N/A":
            sort_key = datetime.max
        else:
            _, parsed = parse_date(row["Bid End Date"])
            sort_key = parsed or datetime.max
        row["_sort_key"] = sort_key
        records.append(row)

    records.sort(key=lambda item: item.get("_sort_key", datetime.max))
    for row in records:
        row.pop("_sort_key", None)

    if output_path is None:
        now = datetime.now()
        filename = f"GeM_Tenders_{now.strftime('%Y%m%d')}_{now.strftime('%H%M')}.xlsx"
        output_path = input_path.parent / filename

    write_workbook(records, errors, output_path)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Transform raw GeM tender JSON/CSV/text into a cleaned XLSX export."
    )
    parser.add_argument("input", type=Path, help="Path to a JSON, CSV, TSV or text file containing raw GeM tender data.")
    parser.add_argument("--output", type=Path, help="Optional output .xlsx path.")
    args = parser.parse_args()

    output = transform_file(args.input, args.output)
    print(f"Saved: {output}")


if __name__ == "__main__":
    main()
