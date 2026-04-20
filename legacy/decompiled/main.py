import os
import re
import pandas as pd
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from selenium.webdriver.common.by import By

from scraper import get_driver, search_keyword, get_bid_cards
from downloader import download_pdf
from pdf_parser import parse_pdf
from extractor import extract_fields, is_it_relevant

# ============================================================================
# CONFIG  — edit these freely
# ============================================================================
KEYWORDS = [
    "Server", "Laptop", "AMC", "CAMC", "FMS",
    "Networking", "Firewall", "HPC", "Workstation",
    "Software Development", "IT Support", "Data Center"
]

BASE_URL          = "https://bidplus.gem.gov.in/all-bids"
CARDS_PER_KEYWORD = 2
MIN_VALUE         = 50000       # reject tenders below this (INR)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SAVE_DIR   = os.path.join(SCRIPT_DIR, "downloads")
OUT_DIR    = os.path.join(SCRIPT_DIR, "output")

# Tender type keywords — used to classify tenders from card text
TENDER_TYPE_MAP = {
    "boq":            ["boq", "bill of quantity", "bill of quantities"],
    "product_custom": ["custom bid", "custom product", "customized"],
    "service":        ["service bid", "service ra", "service contract"],
    "product":        ["product bid", "product ra"],
}


# ============================================================================
# HELPERS
# ============================================================================

def parse_bid_no(text: str) -> str:
    m = re.search(r'Bid\s+No\.?\s*:\s*(GEM/[\w/]+)', text, re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.search(r'\b(GEM/\d{4}/[A-Z]/\d+)\b', text)
    if m:
        return m.group(1).strip()
    return "N/A"


def parse_tender_type(text: str) -> str:
    """
    Infer tender type from card text.
    GeM shows 'Product Bid/RA', 'Service Bid/RA', 'BOQ Bids', etc.
    """
    text_lower = text.lower()
    for ttype, keywords in TENDER_TYPE_MAP.items():
        if any(kw in text_lower for kw in keywords):
            return ttype
    # Check for BOQ marker
    if "boq" in text_lower:
        return "boq"
    # Default — GeM product bids are most common
    return "product"


def parse_card_text(text: str) -> dict:
    lines = [l.strip() for l in text.split("\n") if l.strip()]
    fields = {
        "Bid_No":       parse_bid_no(text),
        "Tender_Type":  parse_tender_type(text),
        "Start_Date":   "N/A",
        "End_Date":     "N/A",
        "Dept_Addr":    "N/A",
        "Items":        "N/A",
        "Quantity":     "N/A",
    }
    for i, line in enumerate(lines):
        if "Items:" in line:
            fields["Items"]      = line.replace("Items:", "").strip()
        elif "Quantity:" in line:
            fields["Quantity"]   = line.replace("Quantity:", "").strip()
        elif "Department Name And Address:" in line:
            fields["Dept_Addr"]  = " | ".join(lines[i+1:i+3])
        elif "Start Date:" in line:
            fields["Start_Date"] = line.replace("Start Date:", "").strip()
        elif "End Date:" in line:
            fields["End_Date"]   = line.replace("End Date:", "").strip()
    return fields


def get_pdf_url(card) -> str:
    if isinstance(card, dict):
        return card.get("pdf_url", "N/A") or "N/A"

    try:
        anchor = card.find_element(By.CSS_SELECTOR, "a.bid_no_hover")
        href   = anchor.get_attribute("href") or ""
        if href and not href.startswith("http"):
            href = "https://bidplus.gem.gov.in" + href
        return href
    except Exception:
        return "N/A"


def value_too_low(bid_value: str) -> bool:
    if bid_value == "N/A":
        return False
    try:
        return int(str(bid_value).replace(",", "")) < MIN_VALUE
    except (ValueError, AttributeError):
        return False


def format_excel_output(writer, df_all, df_it, start_ts, end_ts):
    """Write both sheets with target-style formatting and header date row."""
    date_range = f"Tenders Published: {start_ts} to {end_ts}"

    for sheet_name, df in [("All Tenders", df_all), ("IT Relevant Only", df_it)]:
        # Write date range in row 1, data starts at row 2
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
        ws = writer.sheets[sheet_name]

        # ── Header title row ─────────────────────────────────────────────────
        ws.cell(row=1, column=1, value=date_range)
        ws.cell(row=1, column=1).font = Font(bold=True, size=11)
        ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="1F4E79")
        ws.cell(row=1, column=1).font = Font(bold=True, color="FFFFFF", size=11)

        # Merge header across all columns
        last_col = df.shape[1]
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1,   end_column=last_col
        )

        # ── Column header row styling (row 2) ────────────────────────────────
        header_fill = PatternFill("solid", fgColor="2E75B6")
        for cell in ws[2]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # ── Auto column widths ────────────────────────────────────────────────
        for col_idx, col in enumerate(ws.iter_cols(min_row=2, max_row=ws.max_row), 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(cell.value)) for cell in col if cell.value),
                default=10
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        # ── Wrap text for BOQ and Description columns ─────────────────────────
        wrap_cols = ["Cleaned BOQ", "Description", "Items"]
        for cell in ws[2]:
            if cell.value in wrap_cols:
                col_letter = get_column_letter(cell.column)
                ws.column_dimensions[col_letter].width = 50
                for data_cell in ws[col_letter][2:]:
                    data_cell.alignment = Alignment(wrap_text=True)


# ============================================================================
# MAIN RUNNER
# ============================================================================

def _stop_requested(stop_event) -> bool:
    return stop_event is not None and stop_event.is_set()


def run(stop_event=None):
    driver    = None
    raw_rows  = []
    seen_bids = set()
    run_start = datetime.now()

    print(f"\n🚀 GeM Bot Started — {run_start.strftime('%H:%M:%S')}\n")
    print(f"   Keywords  : {KEYWORDS}")
    print(f"   Cards/KW  : {CARDS_PER_KEYWORD}")
    print(f"   Min Value : ₹{MIN_VALUE:,}\n")

    try:
        if _stop_requested(stop_event):
            print("\nStop requested before start.")
            return

        driver = get_driver()
        driver.get(BASE_URL)

        # Dismiss any popup
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        wait = WebDriverWait(driver, 10)
        try:
            btn = wait.until(EC.element_to_be_clickable((By.XPATH,
                "//button[contains(text(),'Close')] | "
                "//span[contains(text(),'×')] | "
                "//div[contains(@class,'modal-header')]//button"
            )))
            btn.click()
        except Exception:
            pass

        for kw in KEYWORDS:
            if _stop_requested(stop_event):
                print("\nStop requested. Ending run before next keyword.")
                break
            print(f"\n🔍 Keyword: {kw!r}")

            try:
                completed = search_keyword(driver, kw, stop_event=stop_event)
                if not completed:
                    print("   Stop requested during search.")
                    break
                cards = get_bid_cards(driver, limit=CARDS_PER_KEYWORD, stop_event=stop_event)
                print(f"   Found {len(cards)} cards")
            except Exception as e:
                print(f"   [Search Error] {e}")
                continue

            for i, card in enumerate(cards, start=1):
                if _stop_requested(stop_event):
                    print("   Stop requested. Finishing with collected rows only.")
                    break
                try:
                    if isinstance(card, dict):
                        card_text = card.get("text", "")
                    else:
                        parent_card = card.find_element(By.XPATH, "..")
                        card_text = parent_card.text

                    card_fields = parse_card_text(card_text)
                    pdf_url     = get_pdf_url(card)
                    bid_no      = card_fields["Bid_No"]

                    # Deduplication
                    dedup_key = bid_no if bid_no != "N/A" else pdf_url
                    if dedup_key in seen_bids:
                        print(f"   [{i}] SKIP (duplicate): {bid_no}")
                        continue
                    seen_bids.add(dedup_key)

                    print(f"   [{i}] {bid_no}")

                    row = {
                        "Keyword":      kw,
                        "Bid_No":       bid_no,
                        "Tender_Type":  card_fields["Tender_Type"],
                        "Start_Date":   card_fields["Start_Date"],
                        "End_Date":     card_fields["End_Date"],
                        "Dept_Addr":    card_fields["Dept_Addr"],
                        "Items":        card_fields["Items"],
                        "Quantity":     card_fields["Quantity"],
                        "PDF_URL":      pdf_url,
                        # PDF fields defaults
                        "Ministry":     "N/A",
                        "Buyer":        "N/A",
                        "Bid_Value":    "N/A",
                        "EMD":          "N/A",
                        "Delivery":     "N/A",
                        "Category":     "N/A",
                        "Exemption":    "N/A",
                        "Email":        "N/A",
                        "Pincode":      "N/A",
                        "State":        "N/A",
                        "Product_Type": "N/A",
                        "Cleaned_BOQ":  "N/A",
                        "IT_Relevant":  "Pending",
                        "Scraped_At":   datetime.now().strftime("%Y-%m-%d %H:%M"),
                    }

                    # Download + Parse PDF
                    if pdf_url != "N/A":
                        if _stop_requested(stop_event):
                            print("      Stop requested before download")
                            break
                        label    = bid_no if bid_no != "N/A" else f"{kw}_{i}"
                        pdf_path = download_pdf(pdf_url, label, SAVE_DIR)

                        if pdf_path:
                            if _stop_requested(stop_event):
                                print("      Stop requested before PDF parsing")
                                break
                            print(f"      ✦ Parsing {os.path.basename(pdf_path)} …")
                            table_data, full_text = parse_pdf(pdf_path)
                            extracted = extract_fields(
                                table_data, full_text,
                                bid_no=bid_no,
                                dept_addr=row.get("Dept_Addr", ""),
                                items_from_card=row.get("Items", "")
                            )
                            row.update(extracted)
                        else:
                            print(f"      ✗ Download failed")
                    else:
                        print(f"      ⚠ No PDF link found")

                    # IT Relevance filter
                    relevant = is_it_relevant(
                        row.get("Items", ""),
                        row.get("Category", ""),
                        kw
                    )
                    row["IT_Relevant"] = "✅ YES" if relevant else "❌ NO"

                    # Value filter
                    if value_too_low(row.get("Bid_Value", "N/A")):
                        print(f"      ⚠ Skipping — value below ₹{MIN_VALUE:,}")
                        continue

                    raw_rows.append(row)

                except Exception as e:
                    print(f"   [Card Error] {e}")
                    continue

            if _stop_requested(stop_event):
                break

        # ====================================================================
        # BUILD FINAL DATAFRAME — Target format
        # ====================================================================
        if not raw_rows:
            if _stop_requested(stop_event):
                print("\nRun stopped before any data was collected.")
                return
            print("\n⚠ No data collected.")
            return

        run_end = datetime.now()
        start_ts = run_start.strftime("%Y-%m-%d %H:%M")
        end_ts   = run_end.strftime("%Y-%m-%d %H:%M")

        final_rows = []
        it_count   = 0

        for idx, r in enumerate(raw_rows, start=1):
            if r.get("IT_Relevant") == "✅ YES":
                it_count += 1

            # Organisation = Buyer if available, else Dept_Addr first part
            org = r.get("Buyer", "N/A")
            if org == "N/A":
                dept = r.get("Dept_Addr", "N/A")
                if dept != "N/A":
                    org = dept.split("|")[0].strip()

            final_rows.append({
                "S.No":                   idx,
                "Keyword":                r.get("Keyword",       "N/A"),   # ← position 2
                "Tender Reference No":    r.get("Bid_No",        "N/A"),
                "Tender Type":            r.get("Tender_Type",   "N/A"),
                "Published Date":         r.get("Start_Date",    "N/A"),
                "Bid Submission End Date":r.get("End_Date",      "N/A"),
                "Title":                  r.get("Items",         "N/A"),
                "Description":            r.get("Category",      r.get("Items", "N/A")),
                "Cleaned BOQ":            r.get("Cleaned_BOQ",   "N/A"),
                "Organisation":           org,
                "Ministry":               r.get("Ministry",      "N/A"),
                "Tender Value":           r.get("Bid_Value",     "N/A"),   # raw number
                "EMD":                    r.get("EMD",           "N/A"),   # raw number
                "State":                  r.get("State",         "N/A"),
                "Pincode":                r.get("Pincode",       "N/A"),
                "Delivery Period":        r.get("Delivery",      "N/A"),
                "Product Type":           r.get("Product_Type",  "N/A"),
                "Exemption":              r.get("Exemption",     "N/A"),
                "Email":                  r.get("Email",         "N/A"),
                "IT Relevant":            r.get("IT_Relevant",   "N/A"),
                "Quantity":               r.get("Quantity",      "N/A"),
                "Link":                   r.get("PDF_URL",       "N/A"),
                "Scraped Date":           r.get("Scraped_At",    "N/A"),
            })

        df_all = pd.DataFrame(final_rows)
        df_it  = df_all[df_all["IT Relevant"] == "✅ YES"].reset_index(drop=True)
        df_it["S.No"] = range(1, len(df_it) + 1)

        os.makedirs(OUT_DIR, exist_ok=True)
        filename = os.path.join(
            OUT_DIR, f"Tenders_{run_end.strftime('%Y%m%d_%H%M')}.xlsx"
        )

        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            format_excel_output(writer, df_all, df_it, start_ts, end_ts)

        print(f"\n✅ Done!")
        print(f"   Total scraped : {len(df_all)}")
        print(f"   IT Relevant   : {it_count}")
        print(f"   Saved to      : {filename}")
        if _stop_requested(stop_event):
            print("   Status        : Stopped early, partial results saved")

    except Exception as e:
        print(f"\n❌ CRITICAL ERROR: {e}")
        import traceback; traceback.print_exc()

    finally:
        if driver:
            driver.quit()
            print("🧹 Driver closed")


if __name__ == "__main__":
    run()
